import tkinter as tk
from tkinter import filedialog
import sys
import openpyxl
import pandas as pd
import sqlalchemy as sa
from urllib.parse import quote_plus
import os
from openpyxl.utils import get_column_letter

# preprocessing steps:
# source audit file from ingestion team needs to be saved anywhere on computer with suffix yyyymmdd, yyyymmdd should be the date the file was received from platform team
# user variable, redshift user sso value stored in environment variable called rs_user
# pword variable, redshift sso prod password stored in environment variable called rs_finance_prod_sso_pword
# additional python libraries needed if not exist (openpyxl, pandas, sqlalchemy, urllib)

#processing information:
# user will be prompted to select source audit file from computer that was saved with suffix yyyymmdd
# data processing logic will take as input (sso, groname) and add following attributes
# environment, user name, db name, schema name, db role, user role, suggested privilege action
# suggested privilege action will default to Research for any new SSO not in existing lookups in script
# excel file Payroll Database Schema Audit yyyymmdd.xlsx will be produced, formatted, and will open

#ongoing maintenance:
# service_account_lookup_dict should be maintained for new or updated information
# user_role_privilege should be maintained for new or updated information

endpoint = 'odp-fin-prod-etl-redshift.odp.ge-healthcare.net'
user = os.environ['rs_user']
pword = quote_plus(os.environ['rs_finance_prod_sso_pword'])
db_name = 'gehc_data'
port = '5439'
conn_string = f"postgresql+psycopg2://{user}:{pword}@{endpoint}:{port}/{db_name}"
# out_file_test = r'c:\temp\out.xlsx'

service_account_lookup_dict = {
    '502826019': 'DQ FSSO'
    , '504009762': 'Payroll ETL FSSO'
    , '504011355': 'CICD FSSO'
    , 'NP700000489': 'Payroll ETL FSSO'
    , 'NP700000639': 'DQ FSSO'
    , 'NP700002218': 'CICD FSSO'
    , '504009150': 'Payroll/ARCS ETL FSSO'
    , '504011431': 'Payroll REP FSSO'
    , 'NP700000480': 'Payroll/ARCS ETL FSSO'
    , 'NP700000491': 'Payroll REP FSSO'
}

# sso: user_role, suggested_privilege_action
user_role_privilege = {
    '212718864': ['GEHC Architect', 'Keep']
    , '223107692': ['GEHC Architect', 'Keep']
    , '223115412': ['GEHC Architect', 'Keep']
    , '502826019': ['DQ FSSO', 'Keep']
    , '503178675': ['Ingestion (Former)', 'Revoke']
    , '503264401': ['AMS Finance', 'Revoke']
    , '503345949': ['AMS Finance', 'Revoke']
    , '503358774': ['AMS Finance', 'Revoke']
    , '503383638': ['AMS Payroll', 'Keep']
    , '503385972': ['Ingestion', 'Keep']
    , '503392832': ['Ingestion', 'Keep']
    , '504004225': [None, 'Research']
    , '504009762': ['Payroll ETL FSSO', 'Keep']
    , '504011355': ['CICD FSSO', 'Keep']
    , 'NP700000489': ['Payroll ETL FSSO', 'Keep']
    , 'NP700000639': ['DQ FSSO', 'Keep']
    , 'NP700000686': [None, 'Research']
    , 'NP700002218': ['CICD FSSO', 'Keep']
    , '223117412': ['GEHC Program Manager', 'Revoke']
    , '503357526': ['AMS Finance', 'Revoke']
    , '503370023': ['Track (Former)', 'Revoke']
    , '504009150': ['Payroll/ARCS ETL FSSO', 'Keep']
    , '504011431': ['Payroll REP FSSO', 'Keep']
    , '550002371': ['AMS Payroll', 'Keep']
    , '550003015': ['AMS Payroll', 'Keep']
    , '550003875': ['AMS Payroll', 'Keep']
    , '550004871': ['AMS Payroll', 'Keep']
    , '550008650': ['AMS Payroll', 'Keep']
    , '550009432': ['AMS Finance', 'Revoke']
    , '550010641': ['AMS Finance', 'Revoke']
    , 'NP700000480': ['Payroll/ARCS ETL FSSO', 'Keep']
    , 'NP700000491': ['Payroll REP FSSO', 'Keep']
    , '503318601': [None, 'Research']
    , '503395391': ['Track (Former)', 'Revoke']
    , 'NP700007294': [None, 'Research']
    , 'NP700000683': [None, 'Research']
    , '504004019': [None, 'Research']
}


def get_env_sheets(sheet_list_p):
    try:
        return [sheet_name for sheet_name in sheet_list_p if 'etl' in str(sheet_name).lower()]
    except:
        return []


def get_env(sheet_name_p):
    if 'prod' in str(sheet_name_p).lower():
        return 'PROD'
    if 'dev' in str(sheet_name_p).lower():
        return 'NPROD'


def sso_name_lookup(sso_lookup_dict_p, service_account_lookup_dict_p, search_key_p):
    if search_key_p in sso_lookup_dict_p:
        return sso_lookup_dict_p[search_key_p]
    if search_key_p in service_account_lookup_dict_p:
        return service_account_lookup_dict_p[search_key_p]
    else:
        return None


def get_role_from_groname(groname_p):
    last_delim = str(groname_p).rfind('_')
    return groname_p[last_delim + 1:]


def get_schema_from_groname(groname_p):
    start_index = len(db_name + '_')
    last_delim = str(groname_p).rfind('_')
    return groname_p[start_index:last_delim]


def get_user_role(user_role_privilege_p, sso_p):
    if sso_p in user_role_privilege_p:
        return user_role_privilege_p[sso_p][0]
    else:
        return None


def get_suggested_privilege_action(user_role_privilege_p, sso_p):
    if sso_p in user_role_privilege_p:
        return user_role_privilege_p[sso_p][1]
    else:
        return 'Research123'


root = tk.Tk()
root.withdraw()  # hide the main window
file_path = filedialog.askopenfilename(title='select schema audit file')
file_name_without_extension = os.path.basename(file_path).split('.')[0]
date_suffix = file_name_without_extension[-8:]
file_out = r'c:\temp\Payroll Database Schema Audit ' + date_suffix + '.xlsx'

if os.path.exists(file_out):
    os.remove(file_out)

wb_out = openpyxl.Workbook()
wb_out.save(file_out)
wb_out.close()

# check if file name period is valid
if date_suffix.isnumeric() is False or len(date_suffix) != 8:
    print('source file name does not end with date period in format yyyymmdd, process cancelled')
    sys.exit()

if file_path:
    wb = openpyxl.load_workbook(filename=file_path, data_only=True)
    env_sheets = get_env_sheets(wb.sheetnames)
    sheets_to_process = {}
    for sheet_name in env_sheets:
        sheets_to_process[get_env(sheet_name)] = sheet_name
    # get unique user names
    user_names = []
    for k in sheets_to_process:
        sheet_name_iter = sheets_to_process[k]
        ws = wb[sheet_name_iter]
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=1)
            user_names.append(cell.value)
    unique_user_names = set(user_names)
    # exclude NP service accounts
    unique_user_names = [f"'{str(n)}'" for n in unique_user_names if str(n)[0:2] != 'NP']
    user_names_str = ', '.join(unique_user_names)

    # query redshift for sso and names
    engine = sa.create_engine(conn_string)

    sql = f"""
    select emp_sso::varchar(9) as sso
        ,full_name
    from hdl_cmn_view_alpha_ext.prt_persn_d_v
    where emp_sso in ({user_names_str})
    """

    sso_df = pd.read_sql(sql=sql, con=engine)

    sso_lookup_list = sso_df.to_dict(orient='tight')['data']
    sso_lookup_dict = {}
    for x in sso_lookup_list:
        sso_lookup_dict[x[0]] = x[1]

    for k in sheets_to_process:
        print(f'processing {k}: {sheets_to_process[k]}')
        env_iter = k
        ws_iter = wb[sheets_to_process[k]].title
        data_iter_df = pd.read_excel(io=file_path, sheet_name=ws_iter)
        data_iter_df.rename(columns={'usename': 'SSO'}, inplace=True)
        data_iter_df['Environment'] = env_iter
        data_iter_df['SSO'] = data_iter_df['SSO'].astype(str)
        # lookup user name from person table or service account lookup
        data_iter_df['User Name'] = data_iter_df.apply(lambda z: sso_name_lookup(
            sso_lookup_dict, service_account_lookup_dict, z['SSO'])
                                                       , axis=1)
        data_iter_df['DB Name'] = db_name
        data_iter_df['Schema Name'] = data_iter_df.apply(lambda z: get_schema_from_groname(z['groname']), axis=1)
        data_iter_df['DB Role'] = data_iter_df.apply(lambda z: get_role_from_groname(z['groname']), axis=1)
        data_iter_df['User Role'] =  data_iter_df.apply(lambda z: get_user_role(user_role_privilege, z['SSO']), axis=1)
        data_iter_df['Suggested Privilege Action'] = data_iter_df.apply(
            lambda z: get_suggested_privilege_action(user_role_privilege, z['SSO']), axis=1)

        with pd.ExcelWriter(file_out, mode='a') as writer:
            data_iter_df.to_excel(writer, sheet_name=env_iter, index=False)

        # data_iter_df.to_excel(excel_writer=file_out, index=False, sheet_name=env_iter)

else:
    print('no file selected, process cancelled')
    sys.exit()

wb = openpyxl.load_workbook(filename=file_out)

for sheet in wb.sheetnames:
    if str(sheet)[0:5] == 'Sheet':
        wb.remove(wb[sheet])

for sheet in wb.sheetnames:
    ws_iter = wb[sheet]
    for column_cells in ws_iter.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max([len(cell.value or "") for cell in column_cells])
        max_length = max_length + 2
        ws_iter.column_dimensions[column_letter].width = max_length
        ws_iter.freeze_panes = 'A2'

wb.save(file_out)
wb.close()

print('success')
os.startfile(file_out)
