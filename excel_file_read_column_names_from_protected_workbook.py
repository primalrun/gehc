import openpyxl

file_in = r"c:\temp\FA-HealthCare ARCS MBU Template Sep'24.xlsm"

wb = openpyxl.load_workbook(file_in)
ws = wb.active
row_max = ws.max_row

for col in range(1, ws.max_column):
    print(ws.cell(row=8, column=col).value)